from io import BytesIO

from openpyxl import Workbook

from dataline.services.file_parsers.excel_parser import process_excel


def test_process_excel_ignores_blank_header_columns() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(row=2, column=1, value="first_name")
    ws.cell(row=2, column=2, value="last_name")
    ws.cell(row=2, column=3, value="city")

    ws.cell(row=3, column=1, value="Ada")
    ws.cell(row=3, column=2, value="Lovelace")
    ws.cell(row=3, column=3, value="London")

    ws.cell(row=3, column=40, value="ignored")

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    processed_sheets = process_excel(file)

    assert list(processed_sheets["Sheet1"].columns) == ["first_name", "last_name", "city"]
    assert processed_sheets["Sheet1"].to_dict(orient="records") == [
        {"first_name": "Ada", "last_name": "Lovelace", "city": "London"}
    ]
