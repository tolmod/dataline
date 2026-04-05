import logging
import pathlib
import sqlite3
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from pytest import MonkeyPatch

from dataline.config import config
from dataline.models.connection.schema import Connection
from dataline.utils.utils import get_sqlite_dsn

logger = logging.getLogger(__name__)


@pytest.mark.asyncio
async def test_connect_db(client: TestClient) -> None:
    connection_in = {
        "dsn": "sqlite:///test.db",
        "name": "Test",
    }
    response = client.post("/connect", json=connection_in)

    assert response.status_code == 200

    data = response.json()["data"]
    assert data["id"]
    assert data["dsn"] == connection_in["dsn"]
    assert data["name"] == connection_in["name"]
    assert data["dialect"] == "sqlite"
    assert data["database"]
    assert data["is_sample"] is False

    # Delete database after tests
    pathlib.Path("test.db").unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_connect_sample_db(client: TestClient) -> None:
    connection_in = {
        "sample_name": "dvdrental",
        "connection_name": "My DB",
    }
    response = client.post("/connect/sample", json=connection_in)

    assert response.status_code == 200

    data = response.json()["data"]
    assert data["id"]
    assert data["dsn"] is not None
    assert data["dsn"].startswith("sqlite:///")
    assert data["name"] == connection_in["connection_name"]
    assert data["dialect"] == "sqlite"
    assert data["database"]
    assert data["is_sample"] is True

    # Delete database after tests
    file_path = data["dsn"].replace("sqlite:///", "")
    pathlib.Path(file_path).unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_create_sample_db_connection_twice_409(client: TestClient) -> None:
    connection_in = {
        "dsn": get_sqlite_dsn(config.sample_dvdrental_path),
        "name": "Test",
        "is_sample": True,
    }
    response = client.post("/connect", json=connection_in)
    assert response.status_code == 200
    Connection(**response.json()["data"])

    response = client.post("/connect", json=connection_in)
    assert response.status_code == 409


@pytest.mark.asyncio
async def test_get_connections(client: TestClient, dvdrental_connection: Connection) -> None:
    response = client.get("/connections")

    assert response.status_code == 200

    data = response.json()["data"]
    assert data["connections"]
    assert len(data["connections"]) == 1

    connections = data["connections"]
    assert connections[0] == dvdrental_connection.model_dump(mode="json")


@pytest.mark.asyncio
async def test_get_connection(client: TestClient, dvdrental_connection: Connection) -> None:
    response = client.get(f"/connection/{str(dvdrental_connection.id)}")

    assert response.status_code == 200

    data = response.json()["data"]
    assert data == dvdrental_connection.model_dump(mode="json")


@pytest.mark.asyncio
async def test_update_connection(client: TestClient, dvdrental_connection: Connection) -> None:
    update_in = {
        "dsn": "sqlite:///new.db",
        "name": "New name",
    }
    response = client.patch(f"/connection/{str(dvdrental_connection.id)}", json=update_in)

    assert response.status_code == 200

    data = response.json()["data"]
    assert data["connection"]["dsn"] == update_in["dsn"]
    assert data["connection"]["name"] == update_in["name"]

    # Delete database after tests
    pathlib.Path("new.db").unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_delete_connection(client: TestClient, dvdrental_connection: Connection) -> None:
    response = client.delete(f"/connection/{str(dvdrental_connection.id)}")

    assert response.status_code == 200

    # Check if the connection was deleted
    response = client.get("/connections")
    data = response.json()["data"]
    assert len(data["connections"]) == 0


@pytest.mark.asyncio
async def test_connect_excel_file_with_blank_header_columns(
    client: TestClient, monkeypatch: MonkeyPatch, tmp_path: pathlib.Path
) -> None:
    monkeypatch.setattr(config, "data_directory", str(tmp_path))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employees"

    sheet.cell(row=2, column=1, value="first_name")
    sheet.cell(row=2, column=2, value="last_name")
    sheet.cell(row=2, column=3, value="city")

    sheet.cell(row=3, column=1, value="Ada")
    sheet.cell(row=3, column=2, value="Lovelace")
    sheet.cell(row=3, column=3, value="London")
    sheet.cell(row=3, column=40, value="ignored")

    file = BytesIO()
    workbook.save(file)
    file.seek(0)

    response = client.post(
        "/connect/file",
        data={"type": "excel", "name": "Employee Upload"},
        files={
            "file": (
                "employees.xlsx",
                file.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.json()

    data = response.json()["data"]
    file_path = data["dsn"].replace("sqlite:///", "")

    conn = sqlite3.connect(file_path)
    try:
        columns = [row[1] for row in conn.execute("PRAGMA table_info(employees)")]
        rows = conn.execute("SELECT first_name, last_name, city FROM employees").fetchall()
    finally:
        conn.close()
        pathlib.Path(file_path).unlink(missing_ok=True)

    assert columns == ["first_name", "last_name", "city"]
    assert rows == [("Ada", "Lovelace", "London")]
